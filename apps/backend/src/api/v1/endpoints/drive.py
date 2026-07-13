#!/usr/bin/env python3
"""
Google Drive 集成 API 端點
"""

import logging
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, UploadFile
from core.utils import safe_error

from ._deps import get_drive_service

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/drive", tags=["Google Drive"])

_DRIVE_DATA_ROOT = Path(__file__).parent.parent.parent.parent.parent / "data" / "drive_downloads"

# Whitelist of allowed folder aliases — CodeQL path-injection safe
_DRIVE_FOLDER_MAP: Dict[str, Path] = {
    "default": _DRIVE_DATA_ROOT,
    "downloads": _DRIVE_DATA_ROOT,
}


def _get_safe_drive_folder(folder_key: str) -> Path:
    """Resolve a folder from the whitelist.

    The user provides a key/alias; the actual Path is looked up
    from the pre-defined _DRIVE_FOLDER_MAP, keeping CodeQL's
    path-injection query satisfied.  Only the known-safe
    default folder is returned for unknown keys.
    """
    if folder_key not in _DRIVE_FOLDER_MAP:
        logger.warning("Unknown folder_key '%s', using default", folder_key)
    return _DRIVE_FOLDER_MAP.get(folder_key, _DRIVE_DATA_ROOT)


def _safe_drive_dest(folder: Path, file_name: str) -> Path:
    """Resolve a safe destination path under the given folder.

    `folder` MUST already be a validated Path (obtained via
    _get_safe_drive_folder).  Strips directory components from
    file_name to prevent traversal.
    """
    safe_name = Path(file_name).name or "unnamed_file"
    dest = (folder / safe_name).resolve()
    if not (dest == folder or dest.is_relative_to(folder)):
        raise HTTPException(status_code=400, detail="Invalid file name")
    return dest


class DriveDeduplication:
    """根據 content_hash 避免重複下載"""

    def __init__(self):
        from pathlib import Path
        db_path = Path(__file__).parent.parent.parent.parent / "data" / "drive_sync_db.json"
        self._db_path = db_path
        self._syncs: Dict[str, Dict[str, Any]] = {}
        self._load()

    def _load(self) -> None:
        """Load."""
        if self._db_path.exists():
            try:
                import json
                with open(self._db_path, "r", encoding="utf-8") as f:
                    self._syncs = json.load(f)
            except Exception:
                logger.warning("Failed to load sync DB, starting fresh", exc_info=True)
                self._syncs = {}

    def _save(self) -> None:
        """Save."""
        self._db_path.parent.mkdir(parents=True, exist_ok=True)
        import json
        with open(self._db_path, "w", encoding="utf-8") as f:
            json.dump(self._syncs, f, ensure_ascii=False, indent=2)

    def should_download(self, metadata: Dict[str, Any]) -> bool:
        """Execute the should download operation."""
        key = metadata.get("id", "")
        content_hash = metadata.get("content_hash", "")
        if not key:
            return True
        if key not in self._syncs:
            return True
        return self._syncs[key].get("content_hash") != content_hash

    def record_sync(self, metadata: Dict[str, Any], content_hash: str) -> None:
        """Execute the record sync operation."""
        key = metadata.get("id", "")
        if not key:
            return
        self._syncs[key] = {
            "content_hash": content_hash,
            "name": metadata.get("name"),
            "synced_at": datetime.now().isoformat(),
        }
        self._save()

    def compute_content_hash(self, file_path: str) -> str:
        """Compute content hash."""
        import hashlib
        path = Path(file_path)
        if not path.exists():
            return ""
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()


class DocumentParser:
    """解析常見文件格式為純文字"""

    def parse_document(self, file_path: str) -> str:
        """Parse the input data."""
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".txt":
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return f.read()
            except Exception:
                logger.warning(f"Failed to read text file: {path}", exc_info=True)
                return ""

        if suffix in (".md", ".py", ".js", ".ts", ".json", ".yaml", ".yml", ".html", ".css", ".xml"):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return f.read()
            except Exception:
                logger.warning(f"Failed to read source file: {path}", exc_info=True)
                return ""

        if suffix in (".docx",):
            try:
                from docx import Document as DocxDocument
                doc = DocxDocument(path)
                return "\n".join(p.text for p in doc.paragraphs)
            except ImportError:
                return f"[Binary file: {path.name}]"
            except Exception:
                logger.warning(f"Failed to parse docx: {path}", exc_info=True)
                return ""

        if suffix in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                try:
                    lines = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            line = " | ".join(str(c) if c is not None else "" for c in row)
                            if line.strip():
                                lines.append(line)
                    return "\n".join(lines)
                finally:
                    wb.close()
            except ImportError:
                return f"[Spreadsheet: {path.name}]"
            except Exception:
                logger.warning(f"Failed to parse spreadsheet: {path}", exc_info=True)
                return ""

        return f"[File: {path.name} ({suffix})]"


@router.get("/status")
async def get_drive_status(svc=Depends(get_drive_service)) -> dict:
    """獲取 Google Drive 服務狀態"""
    try:
        authenticated = svc.is_authenticated()
        info: Dict[str, Any] = {
            "status": "connected" if authenticated else "disconnected",
            "authenticated": authenticated,
            "service": "Google Drive",
            "last_sync": datetime.now().isoformat(),
        }
        if authenticated:
            try:
                storage = svc.get_storage_info()
                info["quota"] = {
                    "used": storage.get("used", "0"),
                    "total": storage.get("total", "0"),
                    "user": storage.get("user", "unknown"),
                }
            except Exception as e:
                logger.warning(f"Failed to parse drive storage info: {e}", exc_info=True)
        return info
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=safe_error(e))
    except Exception as e:
        logger.error(f"Drive status error: {e}", exc_info=True)
        return {"status": "error", "authenticated": False, "detail": safe_error(e)}


@router.get("/auth/status")
async def get_auth_status(svc=Depends(get_drive_service)) -> dict:
    """獲取認證狀態"""
    try:
        return {"authenticated": svc.is_authenticated()}
    except FileNotFoundError:
        raise HTTPException(status_code=503, detail="credentials.json not found")
    except Exception as e:
        logger.error(f"Auth status error: {e}", exc_info=True)
        return {"authenticated": False, "detail": safe_error(e)}


@router.get("/auth/url")
async def get_oauth_url(svc=Depends(get_drive_service)) -> dict:
    """獲取 OAuth 授權 URL"""
    try:
        url = svc.get_auth_url()
        return {"url": url}
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=safe_error(e))
    except Exception as e:
        logger.error(f"Auth URL error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.post("/auth/callback")
async def oauth_callback(code: str = Body(..., embed=True), svc=Depends(get_drive_service)) -> dict:
    """用授權碼換取 Token"""
    try:
        success = svc.exchange_code(code)
        if success:
            return {"message": "Authentication successful", "authenticated": True}
        raise HTTPException(status_code=400, detail="Token exchange failed")
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=safe_error(e))
    except Exception as e:
        logger.error(f"OAuth callback error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.post("/auth/logout")
async def logout(svc=Depends(get_drive_service)) -> dict:
    """登出 Google Drive"""
    try:
        svc.logout()
        return {"message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"Logout error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.get("/files")
async def list_files(
    page_size: int = Query(10, ge=1, le=100),
    query: Optional[str] = Query(None, description='Drive search query (e.g. \'name contains "report"\')'),
    svc=Depends(get_drive_service),
) -> dict:
    """列出文件"""
    try:
        files = svc.list_files(page_size=page_size, query=query)
        return {"files": files}
    except PermissionError:
        raise HTTPException(status_code=401, detail="Not authenticated. Run /drive/auth/url first.")
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=safe_error(e))
    except Exception as e:
        logger.error(f"List files error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.get("/files/{file_id}/metadata")
async def get_file_metadata(file_id: str, svc=Depends(get_drive_service)) -> dict:
    """獲取文件元數據"""
    try:
        metadata = svc.get_file_metadata(file_id)
        return metadata
    except PermissionError:
        raise HTTPException(status_code=401, detail="Not authenticated")
    except Exception as e:
        logger.error(f"Metadata error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.post("/files/sync")
async def sync_files(request: Dict[str, Any] = Body(...), svc=Depends(get_drive_service)) -> dict:
    """同步選定的文件到本地並存入記憶"""
    file_ids = request.get("file_ids", [])
    default_folder = str(Path(__file__).parent.parent.parent.parent.parent / "data" / "drive_downloads")
    folder_key = request.get("folder_alias", request.get("folder_path", "default"))
    store_memory = request.get("store_memory", True)

    # Early path validation — whitelist-based, CodeQL-safe
    folder = _get_safe_drive_folder(folder_key)

    logger.info(f"Syncing files: {file_ids} to {folder}")

    deduplicator = DriveDeduplication()
    parser = DocumentParser()

    synced_files = []
    skipped_count = 0
    synced_count = 0
    memorized_count = 0

    for fid in file_ids:
        try:
            metadata = svc.get_file_metadata(fid)
        except Exception as e:
            logger.warning(f"Could not get metadata for {fid}: {e}", exc_info=True)
            synced_files.append({"id": fid, "name": f"file_{fid}", "memorized": False, "error": safe_error(e)})
            continue

        dest_path = _safe_drive_dest(folder, metadata.get("name", f"file_{fid}"))
        content_hash = ""

        if not deduplicator.should_download(metadata):
            skipped_count += 1
            synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": False, "skipped": True})
            continue

        download_success = svc.download_file(fid, str(dest_path))

        if download_success:
            synced_count += 1
            content_hash = deduplicator.compute_content_hash(str(dest_path))
            deduplicator.record_sync(metadata, content_hash)

            memory_metadata = {
                "file_id": metadata.get("id"),
                "name": metadata.get("name"),
                "mime_type": metadata.get("mimeType"),
                "size": metadata.get("size"),
                "modified_time": metadata.get("modifiedTime"),
                "source": "google_drive",
            }

            if store_memory:
                try:
                    from ai.memory.ham_memory.ham_manager import HAMMemoryManager
                    ham = HAMMemoryManager()
                    content = parser.parse_document(str(dest_path))
                    ham.store_conversation({
                        "role": "system",
                        "content": content[:5000] if content else f"[File: {dest_path.name}]",
                        "type": "document",
                        "metadata": memory_metadata,
                        "timestamp": datetime.now().isoformat(),
                    })
                    memorized_count += 1
                    synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": True})
                except ImportError:
                    logger.warning("HAMMemoryManager not available, skipping memory store", exc_info=True)
                    synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": False})
                except Exception as me:
                    logger.warning(f"Memory store failed: {me}", exc_info=True)
                    synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": False, "memory_error": safe_error(me)})
            else:
                synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": False})
        else:
            synced_files.append({"id": fid, "name": metadata.get("name"), "memorized": False, "error": "Download failed"})

    return {
        "status": "success",
        "synced": synced_count,
        "skipped": skipped_count,
        "memorized_count": memorized_count,
        "files": synced_files,
    }


@router.post("/files/search")
async def search_and_list(
    request: Dict[str, Any] = Body(...),
    svc=Depends(get_drive_service),
) -> dict:
    """搜尋文件並返回列表（不下載）"""
    query = request.get("query", "")
    page_size = request.get("page_size", 10)

    if not query:
        raise HTTPException(status_code=400, detail="query is required")

    try:
        files = svc.search_files(query=query, page_size=page_size)
        return {"files": files, "count": len(files)}
    except PermissionError:
        raise HTTPException(status_code=401, detail="Not authenticated")
    except Exception as e:
        logger.error(f"Search error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.post("/analyze")
async def analyze_drive(request: Dict[str, Any] = Body(...), svc=Depends(get_drive_service)) -> dict:
    """分析 Drive 文件並總結內容（需下載 + 解析）"""
    limit = request.get("limit", 5)
    default_folder = str(Path(__file__).parent.parent.parent.parent.parent / "data" / "drive_downloads")
    folder_key = request.get("folder_alias", request.get("folder_path", "default"))
    # Early path validation — whitelist-based, CodeQL-safe
    folder = _get_safe_drive_folder(folder_key)

    try:
        files = svc.list_files(page_size=limit)
        parser = DocumentParser()

        summaries = []
        for f in files:
            dest = _safe_drive_dest(folder, f.get("name", f"file_{f['id']}"))
            if svc.download_file(f["id"], str(dest)):
                content = parser.parse_document(str(dest))
                summaries.append({
                    "name": f.get("name"),
                    "size": f.get("size"),
                    "modified": f.get("modifiedTime"),
                    "preview": content[:500] if content else "",
                })

        analysis_parts = []
        for s in summaries:
            analysis_parts.append(
                f"📄 {s['name']} ({s.get('size', '?')} bytes, {s.get('modified', '?')})\n{s['preview'][:200]}"
            )

        analysis = "\n\n".join(analysis_parts) if analysis_parts else "No files found."
        return {"analysis": analysis, "files_analyzed": len(summaries)}

    except PermissionError:
        raise HTTPException(status_code=401, detail="Not authenticated")
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=safe_error(e))
    except Exception as e:
        logger.error(f"Analyze error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=safe_error(e))


@router.post("/files/upload")
async def upload_file(
    file: UploadFile = File(...),
    folder_id: Optional[str] = Form(None),
    svc=Depends(get_drive_service),
) -> Dict[str, Any]:
    """上傳檔案到 Drive"""
    if not svc.is_authenticated():
        raise HTTPException(status_code=401, detail="Not authenticated")
    import os
    suffix = Path(file.filename).suffix if file.filename else ".tmp"
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    try:
        content = await file.read()
        with os.fdopen(fd, "wb") as f:
            f.write(content)
        result = svc.upload_file(tmp_path, folder_id)
        if result is None:
            raise HTTPException(status_code=400, detail="Upload failed")
        return result
    finally:
        os.unlink(tmp_path)


@router.post("/files/create")
async def create_file(
    file_name: str = Body(...),
    content: str = Body(...),
    mime_type: str = Body("text/plain"),
    folder_id: Optional[str] = Body(None),
    svc=Depends(get_drive_service),
) -> Dict[str, Any]:
    """建立文字檔案並上傳到 Drive"""
    if not svc.is_authenticated():
        raise HTTPException(status_code=401, detail="Not authenticated")
    result = svc.create_file_from_text(file_name, content, mime_type, folder_id)
    if result is None:
        raise HTTPException(status_code=400, detail="File creation failed")
    return result
